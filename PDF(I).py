import numpy as np
import pickle
import math
import matplotlib
import matplotlib.pyplot as plt
from scipy import interpolate
import matplotlib.mlab as ml
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import xlsxwriter

I = []
gene = []
R_m = []
R_f = []
wb = load_workbook('D:\Kalyakulina_Anastasiya\Higstogram\PDF.xlsx')
sheet = wb['Sheet1']
i = 0
for i in range(2, 19637):
    I.append(sheet.cell(row=i, column=2).value)
    gene.append(sheet.cell(row=i, column=1).value)
    R_m.append(sheet.cell(row=i, column=3).value)
    R_f.append(sheet.cell(row=i, column=4).value)

wb.close()
I.sort()
R_m.sort()
R_f.sort()
#print(I)
#print(gene)
#print(R_m)
#print(R_f)

'''print(I)
min_elem = min(I)
max_elem = max(I)
i = 0
amount = [0]
new_temp = min(I)
new_age = [min(I)]
l = len(I)
h = (max_elem - min_elem) / 200
for temp in I:
    if temp == new_temp:
        amount[i] += 1
    else:
        amount.append(1)
        new_age.append(temp)
        new_temp = temp
        amount[i] /= (l*h)
        i += 1
amount[i] /= (l*h)
#print(amount)

space = np.linspace(min_elem, max_elem, 201)
x = np.array(new_age)
y = np.array(amount)

f = interpolate.interp1d(x, y, kind='quadratic')
y_new = f(space)
#print(y_new)

plt.plot(space, y_new)
plt.title('PDF')
plt.xlabel('I')
plt.ylabel('P(I)')
#plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\PDF(I).pdf')
plt.show()'''

perce_m = np.percentile(R_m, 90)
perce_f = np.percentile(R_f, 90)
#print(perce_m, perce_f)

new_R_m = []
i = 0
for temp in R_m:
    if temp > perce_m:
        new_R_m.append(temp)

new_R_f = []
for temp in R_f:
    if temp > perce_f:
        new_R_f.append(temp)
'''while R_m[i] < perce_m:
    new_R_m.append(R_m[i])
    i += 1

new_R_f = []
i = 0
while R_f[i] < perce_f:
    new_R_f.append(R_f[i])
    i += 1'''

print(len(new_R_f))
print(len(new_R_m))
'''space = np.linspace(min(R_m), perce_m, 21)
space_new = np.linspace(min(R_m), perce_m, 6)

plt.hist(new_R_m, space, edgecolor='black', color='orange')
plt.xticks(space_new)
plt.title('R2_m distribution')
plt.xlabel('R2_m')
plt.ylabel('Number of values')
#plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\Higstogram_Rm.pdf')
plt.show()

space = np.linspace(min(R_f), perce_f, 21)
space_new = np.linspace(min(R_f), perce_f, 6)

plt.hist(new_R_f, space, edgecolor='black', color='orange')
plt.xticks(space_new)
plt.title('R2_f distribution')
plt.xlabel('R2_f')
plt.ylabel('Number of values')
#plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\Higstogram_Rf.pdf')
plt.show()'''

min_elem = min(R_m)
max_elem = max(R_m)
i = 0
amount = [0]
new_temp = min(R_m)
new_age = [min(R_m)]
l = len(R_m)
h = (max_elem - min_elem) / 20
space = np.linspace(min_elem, max_elem, 21)
print(space)
j = 1
for temp in R_m:
    if j < 21:
        if (temp > new_temp and temp < space[j]) or temp == new_temp:
            amount[i] += 1
        else:
            amount.append(1)
            new_age.append(temp)
            new_temp = space[j]
            j += 1
            amount[i] /= (l*h)
            i += 1
    else:
        if temp > new_temp or temp == new_temp:
            amount[i] += 1
amount[i] /= (l*h)
print(amount)
print(len(amount))
print(len(new_age))

space = np.linspace(min_elem, max_elem, 21)
print(len(space))
x = np.array(new_age)
y = np.array(amount)

f = interpolate.interp1d(x, y, kind='quadratic')
y_new = f(space)

plt.plot(space, y_new)
plt.vlines(perce_m, y.min(), y.max(),  linestyles='--')
plt.xlim(0, max_elem)
plt.ylim(0, max(y_new))
plt.title('PDF')
plt.xlabel('R_m')
plt.ylabel('P(R_m)')
plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\PDF(R_m)_all.pdf')
plt.show()

min_elem = min(R_f)
max_elem = max(R_f)
i = 0
amount = [0]
new_temp = min(R_f)
new_age = [min(R_f)]
l = len(R_f)
h = (max_elem - min_elem) / 20
space = np.linspace(min_elem, max_elem, 21)
j = 1
for temp in R_f:
    if j < 21:
        if (temp > new_temp and temp < space[j]) or temp == new_temp:
            amount[i] += 1
        else:
            amount.append(1)
            new_age.append(temp)
            new_temp = space[j]
            j += 1
            amount[i] /= (l*h)
            i += 1
    else:
        if temp > new_temp or temp == new_temp:
            amount[i] += 1
amount[i] /= (l*h)

space = np.linspace(min_elem, max_elem, 21)
x_o = np.linspace(min_elem, max_elem, 7)
x = np.array(new_age)
y = np.array(amount)

f = interpolate.interp1d(x, y, kind='quadratic')
y_new = f(space)

plt.plot(space, y_new)
plt.vlines(perce_f, y.min(), y.max(),  linestyles='--')
plt.xlim(0, max_elem)
plt.ylim(0, max(y_new))
#plt.xticks(x_o)
plt.title('PDF')
plt.xlabel('R_f')
plt.ylabel('P(R_f)')
plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\PDF(R_f)_all.pdf')
plt.show()