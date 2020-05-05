import numpy as np
import pickle
import math
import matplotlib
import matplotlib.pyplot as plt
from scipy import interpolate
import matplotlib.mlab as ml
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import xlsxwriter
import plotly.graph_objs as go

file = open('D:\Kalyakulina_Anastasiya\OLS\observables.txt', 'r')
age_key = 'age'
gender_key = 'gender'

line = file.readline().rstrip()
line_list = line.split('\t')
age_id = line_list.index(age_key)
gender_id = line_list.index(gender_key)
line_age = []
line_m = []
line_f = []
age_m = []
age_f = []

i = 0
for line in file:
    line_list = line.rstrip().split('\t')
    line_age.append(int(line_list[age_id]))
    if line_list[gender_id] == 'M':
        line_m.append(i)
    else:
        line_f.append(i)
    i += 1
file.close()

for i in line_m:
    age_m.append(line_age[i])
for i in line_f:
    age_f.append(line_age[i])

with open('D:\Kalyakulina_Anastasiya\OLS\gene_row', 'rb') as handle:
    gene_row = pickle.load(handle)

best_list = []
wb = load_workbook('D:\Kalyakulina_Anastasiya\Higstogram\PDF.xlsx')
sheet = wb['Sheet1']
i = 0
for i in range(2, 7):
    best_list.append(sheet.cell(row=i, column=1).value)

wb.close()
#best_list = ['TRIL']
print(best_list)


#name = []
data = np.load('D:\Kalyakulina_Anastasiya\OLS\gene_npz.txt.npz')
betas = data['arr_0']
for gene_name in best_list:
    gene_id = gene_row[gene_name]
    cpg_betas = betas[gene_id]
    betas_m = []
    betas_f = []
    for i in line_m:
        betas_m.append(cpg_betas[i])
    for i in line_f:
        betas_f.append(cpg_betas[i])

    x_min = min(line_age)
    x_max = max(line_age)
    y_min = min(cpg_betas)
    y_max = max(cpg_betas)
    age = 10
    bet = 10
    h_x = (max(line_age) - min(line_age) + 10**(-12))/age
    h_y = (max(cpg_betas) - min(cpg_betas) + 10**(-12))/bet
    eps = 10**(-12)
    mark_x_m = []
    mark_y_m = []
    mark_x_f = []
    mark_y_f = []
    for temp in age_m:
        #id_x = (age*(temp - x_min)/((x_max - x_min) + eps)) + 1
        id_x = ((temp - x_min) / h_x)
        id_x = math.floor(id_x)
        mark_x_m.append(id_x)
    for temp in age_f:
        #id_x = (age*(temp - x_min)/((x_max - x_min) + eps)) + 1
        id_x = ((temp - x_min) / h_x)
        id_x = math.floor(id_x)
        mark_x_f.append(id_x)
    for temp in betas_m:
        #id_y = (bet*(temp - y_min)/((y_max - y_min) + eps)) + 1
        id_y = ((temp - y_min)/h_y)
        id_y = math.floor(id_y)
        mark_y_m.append(id_y)
    for temp in betas_f:
        #id_y = (bet*(temp - y_min)/((y_max - y_min) + eps)) + 1
        id_y = ((temp - y_min) / h_y)
        id_y = math.floor(id_y)
        mark_y_f.append(id_y)

    male = np.zeros((10, 10), dtype=int)
    female = np.zeros((10, 10), dtype=int)

    for temp1, temp2 in zip(mark_x_m, mark_y_m):
        male[temp1, temp2] += 1
    for temp1, temp2 in zip(mark_x_f, mark_y_f):
        female[temp1, temp2] += 1

    i = 0
    j = 0
    new_biggi = np.zeros((10, 10), dtype=int)
    while i < 10:
        j = 0
        while j < 10:
            if (male[i, j] != 0) and (female[i, j] != 0):
                new_biggi[i, j] = min(male[i, j], female[i, j])
            j += 1
        i += 1

    biggi = male + female
    male = male.transpose()
    female = female.transpose()
    biggi = biggi.transpose()
    new_biggi = new_biggi.transpose()

    x = np.linspace(x_min, x_max, 10)
    y = np.linspace(y_min, y_max, 10)

    contour = go.Contour(x=x, y=y, z=new_biggi, contours=dict(coloring='fill', showlines=False), opacity=0.6, line_width=2, colorscale='turbid')
    scatter_f = go.Scatter(x=age_f, y=betas_f, mode='markers', name='female', line=dict(color='red'), opacity=0.5)
    scatter_m = go.Scatter(x=age_m, y=betas_m, mode='markers', name='male', line=dict(color='blue'))
    layout = go.Layout(template='simple_white', plot_bgcolor='rgba(0, 0, 0, 0)', xaxis=dict(title_text='Age', titlefont=dict(size=27), showline=True, ticks='outside', mirror='allticks', showgrid=True, color='black', range=[(x_min - (x_max - x_min)*0.005), (x_max + (x_max - x_min)*0.005)]), yaxis=dict(title_text='Betas', titlefont=dict(size=27), showline=True, ticks='outside', mirror='allticks', showgrid=True, color='black', range=[(y_min - (y_max - y_min)*0.005), (y_max + (y_max - y_min)*0.005)]), font=dict(size=20))
    data = [contour, scatter_m, scatter_f]
    fig = go.Figure(data=data, layout=layout)
    fig.update_layout(title_text=gene_name, width=1200, legend=dict(x=1, y=1.1))
    fig.show()

    '''cm = plt.scatter(age_m, betas_m, alpha=1)
    cf = plt.scatter(age_f, betas_f, edgecolors='red', alpha=0.7)
    cc = plt.contour(x, y, new_biggi, alpha=1)
    plt.colorbar()
    plt.xticks(np.linspace(x_min, x_max, 21))
    plt.title(gene_name)
    plt.xlabel('Age')
    plt.ylabel('Betas')
    plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\{id}_.pdf'.format(id=gene_name))
    plt.show()'''

    '''cc = plt.contourf(x, y, biggi)
    plt.colorbar()
    plt.xticks(np.linspace(x_min, x_max, 21))
    plt.title(gene_name)
    plt.xlabel('Age')
    plt.ylabel('Betas')
    #plt.savefig('D:\Kalyakulina_Anastasiya\Higstogram\{id}.pdf'.format(id=gene_name))
    plt.show()'''

    '''list_contour = [go.Contour(z = male, x = x, y = y, colorscale='blues', opacity=1), go.Contour(z = female, x = x, y = y, colorscale='rdpu', opacity=0.5)]
    fig = go.Figure(list_contour)
    #fig.add_trace(go.Scatter(x=mark_y_m, y=mark_x_m, mode='markers'))
    fig.update_layout(title_text=gene_name)
    fig.show()'''



